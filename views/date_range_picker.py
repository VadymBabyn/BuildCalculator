from PyQt5.QtWidgets import (
    QApplication, QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QCalendarWidget, QMessageBox, QFileDialog
)
from PyQt5.QtCore import QDate
import pandas as pd
from openpyxl.styles import Font, Alignment

from controller.history_controller import HistoryController
from controller.matherial_purchased_controller import MatherialPurchasedController
from controller.payment_controller import PaymentController
from controller.stage_controller import StageController
from controller.sub_stage_controller import SubStageController
class DateRangePicker(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Вибір діапазону дат")
        self.setGeometry(100, 100, 400, 300)

        # Layouts
        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        # Calendar for "від"
        self.label_from = QLabel("Дата від:")
        self.calendar_from = QCalendarWidget()
        self.calendar_from.setGridVisible(True)
        self.calendar_from.setSelectedDate(QDate.currentDate())

        # Calendar for "до"
        self.label_to = QLabel("Дата до:")
        self.calendar_to = QCalendarWidget()
        self.calendar_to.setGridVisible(True)
        self.calendar_to.setSelectedDate(QDate.currentDate())

        # Buttons
        self.buttons_layout = QHBoxLayout()
        self.ok_button = QPushButton("OK")
        self.cancel_button = QPushButton("Скасувати")

        self.buttons_layout.addWidget(self.ok_button)
        self.buttons_layout.addWidget(self.cancel_button)

        # Add widgets to layout
        self.layout.addWidget(self.label_from)
        self.layout.addWidget(self.calendar_from)
        self.layout.addWidget(self.label_to)
        self.layout.addWidget(self.calendar_to)
        self.layout.addLayout(self.buttons_layout)

        # Connect buttons
        self.ok_button.clicked.connect(self.accept)
        self.cancel_button.clicked.connect(self.reject)

        self.history_controller = HistoryController()
        self.stage_controller = StageController()
        self.matherial_purch_controller = MatherialPurchasedController()
        self.payment_controller = PaymentController()
        self.sub_stage_controller = SubStageController()

    def get_selected_dates(self):
        """Повертає обрані дати у форматі (start_date, end_date)."""
        start_date = self.calendar_from.selectedDate().toString("yyyy-MM-dd")
        end_date = self.calendar_to.selectedDate().toString("yyyy-MM-dd")
        return start_date, end_date


    def show_date_picker(self, house_id):
        """Викликає вікно вибору діапазону дат і обробляє результат."""
        dialog = DateRangePicker()
        if dialog.exec_() == QDialog.Accepted:
            start_date, end_date = dialog.get_selected_dates()
            self.generate_report(start_date, end_date, house_id)
            QMessageBox.information(None, "Обраний діапазон", f"Від: {start_date}\nДо: {end_date}")
            return start_date, end_date
        else:
            QMessageBox.information(None, "Скасовано", "Вибір діапазону дат було скасовано.")
            return None, None

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from PyQt5.QtWidgets import QFileDialog, QMessageBox

    def generate_report(self, start_date, end_date, house_id):
        try:
            # Отримання даних у вигляді списку об'єктів History
            history_list = self.history_controller.view_history_by_date(start_date, end_date, house_id)

            # Перевірка, чи є дані
            if not history_list:
                QMessageBox.warning(None, "Звіт", "Дані для вказаного діапазону дат відсутні.")
                return
            stage_name = ""
            sub_stage_name = None
            temp_stage_name = None
            first = True
            # Групування матеріалів за етапами
            materials_by_stage = {}
            for item in history_list:
                if item.payment_id_payment is None:  # Це матеріал
                    if stage_name:
                        temp_stage_name = stage_name
                    temp_sub_stage_name = sub_stage_name
                    stage_name = self.get_stage_name_by_matherial(item.matherial_purchased_id_matherial, None)
                    if stage_name == None:
                        sub_stage_name = self.get_sub_stage_name_by_matherial(item.matherial_purchased_id_matherial, None)
                    if stage_name != "" and stage_name and stage_name not in materials_by_stage:
                        materials_by_stage[stage_name] = []
                    elif sub_stage_name:
                        if temp_sub_stage_name != sub_stage_name:
                            first = True
                        if first:
                            materials_by_stage[temp_stage_name].append({
                                "Назва": "             ",
                                "Од.Вимір": "             ",
                                "Постачальник": "             ",
                                "Кількість": "             ",
                                "Ціна": "             ",
                                "Сума": "             ",
                                "Дата та час": "             ",
                            })
                            materials_by_stage[temp_stage_name].append({
                                "Назва": sub_stage_name,
                                "Од.Вимір": "             ",
                                "Постачальник": "             ",
                                "Кількість": "             ",
                                "Ціна": "             ",
                                "Сума": "             ",
                                "Дата та час": "             ",
                            })
                            materials_by_stage[temp_stage_name].append({
                                "Назва": "             ",
                                "Од.Вимір": "             ",
                                "Постачальник": "             ",
                                "Кількість": "             ",
                                "Ціна": "             ",
                                "Сума": "             ",
                                "Дата та час": "             ",
                            })
                            first = False
                        matherial_info = self.matherial_purch_controller.get_name_and_unit_by_id(
                            item.matherial_purchased_id_matherial)
                        materials_by_stage[temp_stage_name].append({
                            "Назва": matherial_info['matherial_name'],
                            "Од.Вимір": matherial_info['unit'],
                            "Постачальник": item.provider,
                            "Кількість": item.amount,
                            "Ціна": item.price,
                            "Сума": item.sum,
                            "Дата та час": item.time_buy,
                        })
                        continue

                    matherial_info = self.matherial_purch_controller.get_name_and_unit_by_id(item.matherial_purchased_id_matherial)
                    materials_by_stage[stage_name].append({
                        "Назва": matherial_info['matherial_name'],
                        "Од.Вимір": matherial_info['unit'],
                        "Постачальник": item.provider,
                        "Кількість": item.amount,
                        "Ціна": item.price,
                        "Сума": item.sum,
                        "Дата та час": item.time_buy,
                    })
            stage_name = ""
            sub_stage_name = None
            temp_stage_name = None
            first = True
            # Групування послуг за етапами
            services_by_stage = {}
            for item in history_list:
                if item.matherial_purchased_id_matherial is None:  # Це послуга
                    if stage_name:
                        temp_stage_name = stage_name
                    temp_sub_stage_name = sub_stage_name
                    stage_name = self.get_stage_name_by_matherial(None, item.payment_id_payment)
                    if stage_name == None:
                        sub_stage_name = self.get_sub_stage_name_by_matherial(None, item.payment_id_payment)
                    if stage_name != "" and stage_name and stage_name not in services_by_stage:
                        services_by_stage[stage_name] = []
                    elif sub_stage_name:
                        if temp_sub_stage_name != sub_stage_name:
                            first = True
                        if first:
                            services_by_stage[temp_stage_name].append({
                                "Назва": "             ",
                                "Од.Вимір": "             ",
                                "Постачальник": "             ",
                                "Кількість": "             ",
                                "Ціна": "             ",
                                "Сума": "             ",
                                "Дата та час": "             ",
                            })
                            services_by_stage[temp_stage_name].append({
                                "Назва": sub_stage_name,
                                "Од.Вимір": "             ",
                                "Постачальник": "             ",
                                "Кількість": "             ",
                                "Ціна": "             ",
                                "Сума": "             ",
                                "Дата та час": "             ",
                            })
                            services_by_stage[temp_stage_name].append({
                                "Назва": "             ",
                                "Од.Вимір": "             ",
                                "Постачальник": "             ",
                                "Кількість": "             ",
                                "Ціна": "             ",
                                "Сума": "             ",
                                "Дата та час": "             ",
                            })
                            first = False
                        payment_info = self.payment_controller.get_name_and_unit_by_id(item.payment_id_payment)
                        services_by_stage[temp_stage_name].append({
                            "Назва": payment_info['matherial_name'],
                            "Од.Вимір": payment_info['unit'],
                            "Отримувач": item.provider,
                            "Кількість": item.amount,
                            "Ціна": item.price,
                            "Сума": item.sum,
                            "Дата та час": item.time_buy,
                        })
                        continue
                    payment_info = self.payment_controller.get_name_and_unit_by_id(item.payment_id_payment)
                    if stage_name not in services_by_stage:
                        services_by_stage[stage_name] = []
                    services_by_stage[stage_name].append({
                        "Назва": payment_info['matherial_name'],
                        "Од.Вимір": payment_info['unit'],
                        "Отримувач": item.provider,
                        "Кількість": item.amount,
                        "Ціна": item.price,
                        "Сума": item.sum,
                        "Дата та час": item.time_buy,
                    })

            # Відображення діалогового вікна для вибору місця збереження файлу
            options = QFileDialog.Options()
            file_path, _ = QFileDialog.getSaveFileName(
                None,
                "Зберегти звіт",
                f"review_{start_date}_to_{end_date}.xlsx",
                "Excel Files (*.xlsx)",
                options=options
            )

            if not file_path:  # Якщо користувач натиснув "Скасувати"
                return

            # Створення Excel-файлу з розділенням по листах
            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                # Додавання матеріалів на листи за етапами
                for stage_name, materials in materials_by_stage.items():
                    df = pd.DataFrame(materials)
                    if not df.empty:
                        df.to_excel(writer, sheet_name=f"Матеріали ({stage_name})", index=False)

                # Додавання послуг на листи за етапами
                for stage_name, services in services_by_stage.items():
                    df = pd.DataFrame(services)
                    if not df.empty:
                        df.to_excel(writer, sheet_name=f"Послуги ({stage_name})", index=False)

            # Налаштування стилю Excel-файлу
            from openpyxl import load_workbook

            workbook = load_workbook(file_path)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                # Встановлення ширини стовпців
                for column_cells in sheet.columns:
                    max_length = 0
                    column_letter = column_cells[0].column_letter  # Отримати літеру стовпця
                    for cell in column_cells:
                        try:
                            if cell.value:  # Перевірка, чи є значення
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = max_length + 10  # Додатковий відступ
                    sheet.column_dimensions[column_letter].width = adjusted_width

                # Встановлення стилю шрифту для заголовків
                for cell in sheet[1]:  # Перший рядок - заголовки
                    cell.font = Font(bold=True, size=16)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.row != 1:
                            cell.font = Font(size=16)
            workbook.save(file_path)

            QMessageBox.information(None, "Звіт", f"Звіт успішно збережено у файл: {file_path}")
        except Exception as e:
            QMessageBox.critical(None, "Помилка", f"Не вдалося створити звіт: {e}")

    def get_sub_stage_name_by_matherial(self, matherial_purch_id=None, payment_id=None):
        # Приклад отримання назви етапу через контролер
        if matherial_purch_id:
            id_sub_stage = self.matherial_purch_controller.take_sub_stage_id_by_matherial_id(matherial_purch_id, payment_id)
        else:
            id_sub_stage = self.matherial_purch_controller.take_sub_stage_id_by_matherial_id(matherial_purch_id, payment_id)
        return self.sub_stage_controller.get_sub_stage_name_by_id(id_sub_stage)

    def get_stage_name_by_matherial(self, matherial_purch_id=None, payment_id=None):
        # Приклад отримання назви етапу через контролер
        if matherial_purch_id:
            id_stage = self.matherial_purch_controller.take_stage_id_by_matherial_id(matherial_purch_id, payment_id)
        else:
            id_stage = self.matherial_purch_controller.take_stage_id_by_matherial_id(matherial_purch_id, payment_id)
        return self.stage_controller.get_stage_name_by_id(id_stage)


# Приклад підключення до кнопки
if __name__ == "__main__":
    import sys

    app = QApplication(sys.argv)

    sys.exit(app.exec_())
